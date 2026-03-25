import gc
import logging
import os
from datetime import date as dt_date

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from domain.import_policy import ImportPolicy
from domain.records import MandatoryExpenseRecord, Record
from domain.reports import Report
from services.import_parser import parse_transfer_row
from utils.csv_utils import (
    _restore_missing_transfers,
    _validate_transfer_integrity,
)
from utils.import_core import (
    ImportSummary,
    norm_key,
    parse_import_row,
    safe_type,
)
from utils.money import to_money_float
from utils.tabular_utils import (
    mandatory_expense_export_rows,
    record_export_rows,
    report_record_type_label,
    resolve_get_rate,
)

logger = logging.getLogger(__name__)
MAX_IMPORT_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
MAX_IMPORT_ROWS = 200_000

DATA_HEADERS = [
    "date",
    "type",
    "wallet_id",
    "category",
    "amount_original",
    "currency",
    "rate_at_operation",
    "amount_kzt",
    "description",
    "period",
    "transfer_id",
    "from_wallet_id",
    "to_wallet_id",
]
MANDATORY_HEADERS = [
    "type",
    "date",
    "category",
    "amount_original",
    "currency",
    "rate_at_operation",
    "amount_kzt",
    "description",
    "period",
]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SUBTOTAL_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
TOTAL_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _style_title_row(ws, row_idx: int, *, columns: int) -> None:
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=columns)
    cell = ws.cell(row=row_idx, column=1)
    cell.font = Font(bold=True, size=14, color="1F1F1F")
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _style_header_row(ws, row_idx: int, *, center: bool = False) -> None:
    for cell in ws[row_idx]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="center",
        )


def _style_data_row(ws, row_idx: int, *, amount_columns: tuple[int, ...] = ()) -> None:
    for cell in ws[row_idx]:
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for column_idx in amount_columns:
        ws.cell(row=row_idx, column=column_idx).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=column_idx).alignment = Alignment(
            horizontal="right",
            vertical="center",
        )


def _style_total_row(
    ws, row_idx: int, *, fill: PatternFill, amount_columns: tuple[int, ...]
) -> None:
    for cell in ws[row_idx]:
        cell.font = Font(bold=True, color="1F1F1F")
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for column_idx in amount_columns:
        ws.cell(row=row_idx, column=column_idx).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=column_idx).alignment = Alignment(
            horizontal="right",
            vertical="center",
        )


def _set_auto_width(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), len(value))
    for column_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(column_idx)].width = min(max(width + 2, 12), 32)


def _safe_str(value):
    return "" if value is None else str(value)


def _should_add_by_category_sheet(report: Report, groups: dict[str, Report]) -> bool:
    if len(groups) > 1:
        return True
    if len(groups) != 1:
        return False
    only_subreport = next(iter(groups.values()))
    return len(list(only_subreport.records())) < len(list(report.records()))


def report_to_xlsx(report: Report, filepath: str) -> None:
    """Export report view (fixed amounts) to XLSX. Read-only format."""
    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "Report"
        ws.append([report.statement_title, "", "", ""])
        _style_title_row(ws, 1, columns=4)
        ws.append(["Date", "Type", "Category", "Amount (KZT)"])
        _style_header_row(ws, 2)
        ws.append(["", "", "", "Fixed amounts by operation-time FX rates"])
        ws["D3"].alignment = Alignment(horizontal="right", vertical="center")
        ws["D3"].font = Font(italic=True, color="666666")
        ws.freeze_panes = "A3"

    if (getattr(report, "initial_balance", 0) != 0 or report.is_opening_balance) and ws is not None:
        ws.append(["", report.balance_label, "", report.initial_balance])
        _style_total_row(ws, ws.max_row, fill=SECTION_FILL, amount_columns=(4,))

    for record in sorted(
        report.records(),
        key=lambda r: (0, r.date) if isinstance(r.date, dt_date) else (1, dt_date.max),
    ):
        if ws is not None:
            record_date = (
                record.date.isoformat() if isinstance(record.date, dt_date) else record.date
            )
            ws.append(
                [
                    record_date,
                    report_record_type_label(record),
                    record.category,
                    record.amount_kzt,
                ]
            )
            _style_data_row(ws, ws.max_row, amount_columns=(4,))

    total = report.total_fixed()
    records_total = sum(r.signed_amount_kzt() for r in report.records())
    if ws is not None:
        ws.append(["SUBTOTAL", "", "", records_total])
        _style_total_row(ws, ws.max_row, fill=SUBTOTAL_FILL, amount_columns=(4,))
        ws.append(["FINAL BALANCE", "", "", total])
        _style_total_row(ws, ws.max_row, fill=TOTAL_FILL, amount_columns=(4,))
        ws.auto_filter.ref = f"A2:D{ws.max_row}"
        _set_auto_width(ws)

    summary_year, monthly_rows = report.monthly_income_expense_rows()
    summary_ws = wb.create_sheet("Yearly Report")
    if summary_ws is not None:
        summary_ws.append([f"Month ({summary_year})", "Income (KZT)", "Expense (KZT)"])
        _style_header_row(summary_ws, 1, center=True)
        summary_ws.freeze_panes = "A2"
        total_income = 0.0
        total_expense = 0.0
        for month_label, income, expense in monthly_rows:
            total_income += income
            total_expense += expense
            summary_ws.append([month_label, income, expense])
            _style_data_row(summary_ws, summary_ws.max_row, amount_columns=(2, 3))
        summary_ws.append(["TOTAL", total_income, total_expense])
        _style_total_row(summary_ws, summary_ws.max_row, fill=TOTAL_FILL, amount_columns=(2, 3))
        summary_ws.auto_filter.ref = f"A1:C{summary_ws.max_row}"
        _set_auto_width(summary_ws)

    try:
        groups = report.grouped_by_category()
    except Exception:
        groups = {}

    if _should_add_by_category_sheet(report, groups):
        bycat_ws = wb.create_sheet(title="By Category", index=1)
        for category, subreport in sorted(groups.items(), key=lambda x: x[0] or ""):
            bycat_ws.append([f"Category: {category}"])
            category_row = bycat_ws.max_row
            bycat_ws.merge_cells(
                start_row=category_row,
                start_column=1,
                end_row=category_row,
                end_column=3,
            )
            bycat_ws.cell(row=category_row, column=1).font = Font(bold=True, color="1F1F1F")
            bycat_ws.cell(row=category_row, column=1).fill = SECTION_FILL
            bycat_ws.cell(row=category_row, column=1).border = THIN_BORDER
            bycat_ws.append(["Date", "Type", "Amount (KZT)"])
            _style_header_row(bycat_ws, bycat_ws.max_row)
            records_total = 0.0
            for r in sorted(
                subreport.records(),
                key=lambda rr: (0, rr.date) if isinstance(rr.date, dt_date) else (1, dt_date.max),
            ):
                amt = getattr(r, "amount", 0.0)
                records_total += (
                    getattr(r, "amount", 0.0) if getattr(r, "amount", None) is not None else 0.0
                )
                display_date = getattr(r, "date", "")
                if isinstance(display_date, dt_date):
                    display_date = display_date.isoformat()
                bycat_ws.append([display_date, report_record_type_label(r), abs(amt)])
                _style_data_row(bycat_ws, bycat_ws.max_row, amount_columns=(3,))
            bycat_ws.append(["SUBTOTAL", "", abs(records_total)])
            _style_total_row(bycat_ws, bycat_ws.max_row, fill=SUBTOTAL_FILL, amount_columns=(3,))
            bycat_ws.append([""])
        bycat_ws.freeze_panes = "A2"
        _set_auto_width(bycat_ws)

    os.makedirs(os.path.dirname(filepath), exist_ok=True) if os.path.dirname(filepath) else None
    wb.save(filepath)
    try:
        wb.close()
    except Exception:
        pass
    gc.collect()


def grouped_report_to_xlsx(
    statement_title: str,
    grouped_rows: list[tuple[str, int, float]],
    filepath: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "Grouped Report"
        ws.append([statement_title, "", ""])
        _style_title_row(ws, 1, columns=3)
        ws.append(["Category", "Operations", "Total (KZT)"])
        _style_header_row(ws, 2)
        ws.append(["", "", "Grouped category totals"])
        ws["C3"].alignment = Alignment(horizontal="right", vertical="center")
        ws["C3"].font = Font(italic=True, color="666666")
        ws.freeze_panes = "A3"

        total_kzt = 0.0
        for category, operations_count, amount_kzt in grouped_rows:
            total_kzt += float(amount_kzt)
            ws.append([category, int(operations_count), float(amount_kzt)])
            _style_data_row(ws, ws.max_row, amount_columns=(3,))

        ws.append(["TOTAL", "", total_kzt])
        _style_total_row(ws, ws.max_row, fill=TOTAL_FILL, amount_columns=(3,))
        ws.auto_filter.ref = f"A2:C{ws.max_row}"
        _set_auto_width(ws)

    os.makedirs(os.path.dirname(filepath), exist_ok=True) if os.path.dirname(filepath) else None
    wb.save(filepath)
    try:
        wb.close()
    except Exception:
        pass
    gc.collect()


def report_from_xlsx(filepath: str) -> Report:
    records, initial_balance, _ = import_records_from_xlsx(filepath, ImportPolicy.LEGACY)
    return Report(records, initial_balance)


def export_records_to_xlsx(
    records: list[Record],
    filepath: str,
    initial_balance: float = 0.0,
    *,
    transfers=None,
) -> None:
    del initial_balance  # legacy argument, kept for compatibility

    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "Data"
        ws.append(DATA_HEADERS)
        _style_header_row(ws, 1)
        ws.freeze_panes = "A2"

    for row in record_export_rows(records, transfers=list(transfers or [])):
        if ws is not None:
            ws.append([row.get(header, "") for header in DATA_HEADERS])
            _style_data_row(ws, ws.max_row, amount_columns=(5, 7, 8))

    if ws is not None and ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:M{ws.max_row}"
        _set_auto_width(ws)

    os.makedirs(os.path.dirname(filepath), exist_ok=True) if os.path.dirname(filepath) else None
    wb.save(filepath)
    try:
        wb.close()
    except Exception:
        pass
    gc.collect()


def import_records_from_xlsx(
    filepath: str,
    policy: ImportPolicy = ImportPolicy.FULL_BACKUP,
    currency_service=None,
    wallet_ids: set[int] | None = None,
    existing_initial_balance: float = 0.0,
) -> tuple[list[Record], float, ImportSummary]:
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"XLSX file not found: {filepath}")
    if os.path.getsize(filepath) > MAX_IMPORT_FILE_SIZE:
        raise ValueError(f"XLSX file is too large: {os.path.getsize(filepath)} bytes")

    wb = load_workbook(filepath, data_only=True, read_only=True)
    try:
        if not wb.worksheets:
            return [], 0.0, (0, 0, [])

        ws = wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        first_row = next(rows_iter, None)
        if first_row is None:
            return [], 0.0, (0, 0, [])

        header_offset = 0
        header_row = first_row
        if first_row and _safe_str(first_row[0]).strip().startswith("Transaction statement"):
            header_row = next(rows_iter, None)
            header_offset = 1
        if header_row is None:
            return [], 0.0, (0, 0, [])

        headers = [norm_key(_safe_str(h)) for h in header_row]
        records: list[Record] = []
        initial_balance = to_money_float(existing_initial_balance)
        errors: list[str] = []
        skipped = 0
        imported = 0

        transfers = {}
        next_transfer_id = 1

        is_report_xlsx = {"date", "type", "category", "amount_(kzt)"}.issubset(set(headers))
        get_rate = None
        if policy == ImportPolicy.CURRENT_RATE:
            get_rate = resolve_get_rate(currency_service)
        seen_rows = 0
        seen_initial_balance = False

        for idx, row in enumerate(rows_iter, start=header_offset + 2):
            seen_rows += 1
            if seen_rows > MAX_IMPORT_ROWS:
                raise ValueError(f"XLSX import exceeded row limit ({MAX_IMPORT_ROWS})")
            if not row or all(cell is None for cell in row):
                continue

            raw = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}

            if is_report_xlsx:
                date_value = _safe_str(raw.get("date", "")).strip()
                if date_value.upper() in {"SUBTOTAL", "FINAL_BALANCE", "FINAL BALANCE"}:
                    continue
                if date_value == "" and norm_key(_safe_str(raw.get("type", "")).strip()) in {
                    "initial_balance",
                    "opening_balance",
                }:
                    raw["type"] = "initial_balance"
                    raw["amount_original"] = raw.get("amount_(kzt)")
                else:
                    raw["amount"] = raw.get("amount_(kzt)")

            row_type = safe_type(_safe_str(raw.get("type", "")).lower())
            if row_type == "transfer":
                row_lc = {norm_key(str(k)): str(v) if v is not None else "" for k, v in raw.items()}
                parsed_records, transfer, next_transfer_id, error = parse_transfer_row(
                    row_lc=row_lc,
                    row_label=f"row {idx}",
                    policy=policy,
                    get_rate=get_rate,
                    next_transfer_id=next_transfer_id,
                    wallet_ids=wallet_ids,
                )
                if error:
                    skipped += 1
                    errors.append(error)
                    logger.warning("XLSX import skipped %s", error)
                    continue
                if transfer is not None and parsed_records is not None:
                    if transfer.id in transfers:
                        skipped += 1
                        errors.append(f"row {idx}: duplicate transfer_id #{transfer.id}")
                        continue
                    transfers[transfer.id] = transfer
                    records.extend(parsed_records)
                    imported += 1
                continue

            record, parsed_balance, error = parse_import_row(
                raw,
                row_label=f"row {idx}",
                policy=policy,
                get_rate=get_rate,
                mandatory_only=False,
            )
            if error:
                skipped += 1
                errors.append(error)
                logger.warning("XLSX import skipped %s", error)
                continue
            if parsed_balance is not None:
                if seen_initial_balance:
                    skipped += 1
                    errors.append(f"row {idx}: duplicate initial_balance row")
                    logger.warning("XLSX import skipped row %s: duplicate initial_balance row", idx)
                    continue
                seen_initial_balance = True
                initial_balance = parsed_balance
                continue
            if record is None:
                continue
            if wallet_ids is not None and record.wallet_id not in wallet_ids:
                skipped += 1
                errors.append(f"row {idx}: wallet not found ({record.wallet_id})")
                continue
            imported += 1
            records.append(record)
            if record.transfer_id is not None:
                next_transfer_id = max(next_transfer_id, int(record.transfer_id) + 1)

        _restore_missing_transfers(records, transfers)
        integrity_errors = _validate_transfer_integrity(records, transfers, wallet_ids)
        if integrity_errors:
            skipped += len(integrity_errors)
            errors.extend(integrity_errors)

        logger.info(
            "XLSX import completed: imported=%s skipped=%s file=%s",
            imported,
            skipped,
            filepath,
        )
        return records, initial_balance, (imported, skipped, errors)
    finally:
        try:
            wb.close()
        except Exception:
            pass
        gc.collect()


def export_mandatory_expenses_to_xlsx(
    expenses: list[MandatoryExpenseRecord], filepath: str
) -> None:
    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "Mandatory"
        ws.append(MANDATORY_HEADERS)
        _style_header_row(ws, 1)
        ws.freeze_panes = "A2"

    for row in mandatory_expense_export_rows(expenses):
        if ws is not None:
            ws.append([row.get(header, "") for header in MANDATORY_HEADERS])
            _style_data_row(ws, ws.max_row, amount_columns=(4, 6, 7))

    if ws is not None and ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:I{ws.max_row}"
        _set_auto_width(ws)

    os.makedirs(os.path.dirname(filepath), exist_ok=True) if os.path.dirname(filepath) else None
    wb.save(filepath)
    try:
        wb.close()
    except Exception:
        pass
    gc.collect()


def import_mandatory_expenses_from_xlsx(
    filepath: str,
    policy: ImportPolicy = ImportPolicy.FULL_BACKUP,
    currency_service=None,
) -> tuple[list[MandatoryExpenseRecord], ImportSummary]:
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"XLSX file not found: {filepath}")
    if os.path.getsize(filepath) > MAX_IMPORT_FILE_SIZE:
        raise ValueError(f"XLSX file is too large: {os.path.getsize(filepath)} bytes")

    wb = load_workbook(filepath, data_only=True, read_only=True)
    try:
        if not wb.worksheets:
            return [], (0, 0, [])

        ws = wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            return [], (0, 0, [])

        headers = [norm_key(_safe_str(h)) for h in header_row]
        expenses: list[MandatoryExpenseRecord] = []
        errors: list[str] = []
        skipped = 0
        imported = 0

        get_rate = None
        if policy == ImportPolicy.CURRENT_RATE:
            get_rate = resolve_get_rate(currency_service)
        seen_rows = 0

        for idx, row in enumerate(rows_iter, start=2):
            seen_rows += 1
            if seen_rows > MAX_IMPORT_ROWS:
                raise ValueError(f"XLSX import exceeded row limit ({MAX_IMPORT_ROWS})")
            if not row or all(cell is None for cell in row):
                continue
            raw = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}

            record, _, error = parse_import_row(
                raw,
                row_label=f"row {idx}",
                policy=policy,
                get_rate=get_rate,
                mandatory_only=True,
            )
            if error:
                skipped += 1
                errors.append(error)
                logger.warning("Mandatory XLSX import skipped %s", error)
                continue
            if isinstance(record, MandatoryExpenseRecord):
                imported += 1
                expenses.append(record)

        logger.info(
            "Mandatory XLSX import completed: imported=%s skipped=%s file=%s",
            imported,
            skipped,
            filepath,
        )
        return expenses, (imported, skipped, errors)
    finally:
        try:
            wb.close()
        except Exception:
            pass
        gc.collect()

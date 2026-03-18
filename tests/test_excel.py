import os
import tempfile
import time

from openpyxl import load_workbook

from domain.records import (
    ExpenseRecord,
    IncomeRecord,
    MandatoryExpenseRecord,
)
from domain.reports import Report
from utils.csv_utils import (
    export_mandatory_expenses_to_csv,
    import_mandatory_expenses_from_csv,
)
from utils.excel_utils import (
    export_mandatory_expenses_to_xlsx,
    import_mandatory_expenses_from_xlsx,
    report_from_xlsx,
    report_to_xlsx,
)


def test_report_xlsx_roundtrip():
    records = [
        IncomeRecord(date="2025-01-01", _amount_init=100.0, category="Salary"),
        ExpenseRecord(date="2025-01-02", _amount_init=30.0, category="Food"),
    ]
    report = Report(records, initial_balance=50.0)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    try:
        report_to_xlsx(report, tmp_path)
        wb = load_workbook(tmp_path, data_only=True)
        try:
            report_ws = wb["Report"]
            assert report_ws.cell(1, 1).value == "Transaction statement"
            assert report_ws.cell(4, 2).value == "Initial balance"
            assert "Yearly Report" in wb.sheetnames
            summary_ws = wb["Yearly Report"]
            assert summary_ws.cell(1, 1).value == "Month (2025)"
        finally:
            wb.close()
        imported = report_from_xlsx(tmp_path)
        assert len(imported.records()) == 2
        assert abs(imported._initial_balance - 50.0) < 1e-6
        assert abs(imported.total() - report.total()) < 1e-6
    finally:
        os.unlink(tmp_path)


def test_report_xlsx_uses_opening_balance_label_for_filtered_report():
    records = [
        IncomeRecord(date="2024-12-31", _amount_init=20.0, category="Old"),
        IncomeRecord(date="2025-01-01", _amount_init=100.0, category="Salary"),
        ExpenseRecord(date="2025-01-02", _amount_init=30.0, category="Food"),
    ]
    report = Report(records, initial_balance=50.0).filter_by_period("2025")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    try:
        report_to_xlsx(report, tmp_path)
        wb = load_workbook(tmp_path, data_only=True)
        try:
            ws = wb["Report"]
            assert ws.cell(1, 1).value == "Transaction statement (2025-01-01 - 2025-12-31)"
            assert ws.cell(4, 2).value == "Opening balance"
            assert ws.cell(4, 4).value == "70.00"
        finally:
            wb.close()
    finally:
        for _ in range(5):
            try:
                os.unlink(tmp_path)
                break
            except PermissionError:
                time.sleep(0.1)


def test_mandatory_xlsx_roundtrip():
    expenses = [
        MandatoryExpenseRecord(
            date="2026-03-09",
            _amount_init=10.0,
            category="Sub",
            description="d1",
            period="monthly",
            auto_pay=True,
        ),
        MandatoryExpenseRecord(
            date="",
            _amount_init=20.5,
            category="Svc",
            description="d2",
            period="yearly",
        ),
    ]

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    try:
        export_mandatory_expenses_to_xlsx(expenses, tmp_path)
        imported, _ = import_mandatory_expenses_from_xlsx(tmp_path)
        assert len(imported) == 2
        assert imported[0].amount == 10.0
        assert str(imported[0].date) == "2026-03-09"
        assert imported[0].auto_pay is True
        assert imported[1].period == "yearly"
    finally:
        os.unlink(tmp_path)


def test_mandatory_csv_roundtrip():
    expenses = [
        MandatoryExpenseRecord(
            date="2026-03-21",
            _amount_init=5.0,
            category="A",
            description="x",
            period="daily",
            auto_pay=True,
        ),
    ]
    with tempfile.NamedTemporaryFile(delete=False, suffix=".csv", mode="w", newline="") as tmp:
        tmp_path = tmp.name
    try:
        export_mandatory_expenses_to_csv(expenses, tmp_path)
        imported, _ = import_mandatory_expenses_from_csv(tmp_path)
        assert len(imported) == 1
        assert imported[0].amount == 5.0
        assert str(imported[0].date) == "2026-03-21"
        assert imported[0].auto_pay is True
        assert imported[0].period == "daily"
    finally:
        os.unlink(tmp_path)


def test_xlsx_export_grouped_drill_down():
    """Export a category‑filtered report to XLSX (simulating grouped drill‑down)."""
    import os
    import tempfile

    from openpyxl import load_workbook

    from domain.records import ExpenseRecord, IncomeRecord
    from domain.reports import Report
    from utils.excel_utils import report_to_xlsx

    records = [
        IncomeRecord(date="2025-01-01", _amount_init=100.0, category="Salary"),
        ExpenseRecord(date="2025-01-02", _amount_init=30.0, category="Food"),
        IncomeRecord(date="2025-01-03", _amount_init=50.0, category="Salary"),
        ExpenseRecord(date="2025-01-04", _amount_init=20.0, category="Food"),
    ]
    report = Report(records, initial_balance=200.0)
    filtered = report.filter_by_category("Salary")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    try:
        report_to_xlsx(filtered, tmp_path)
        wb = load_workbook(tmp_path, data_only=True)
        try:
            ws = wb["Report"]
            # Check title
            assert ws.cell(1, 1).value == "Transaction statement"
            # Check that there is no initial balance row (because filter_by_category resets it)
            # The first data row should be the first Salary record
            data_rows = []
            for row in ws.iter_rows(min_row=4, max_col=4, values_only=True):
                if row[0] and isinstance(row[0], str) and row[0].startswith("2025"):
                    data_rows.append(row)
            assert len(data_rows) == 2
            assert data_rows[0][2] == "Salary"
            assert data_rows[0][3] == "100.00"
            assert data_rows[1][2] == "Salary"
            assert data_rows[1][3] == "50.00"
            # Find SUBTOTAL row
            subtotal_found = False
            for row in ws.iter_rows(min_row=1, max_col=4, values_only=True):
                if row[0] == "SUBTOTAL":
                    assert row[3] == "150.00"
                    subtotal_found = True
                    break
            assert subtotal_found
            # Find FINAL BALANCE row
            final_found = False
            for row in ws.iter_rows(min_row=1, max_col=4, values_only=True):
                if row[0] == "FINAL BALANCE":
                    assert row[3] == "150.00"
                    final_found = True
                    break
            assert final_found
        finally:
            wb.close()
    finally:
        for _ in range(5):
            try:
                os.unlink(tmp_path)
                break
            except PermissionError:
                time.sleep(0.1)

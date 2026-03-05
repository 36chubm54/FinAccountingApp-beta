from __future__ import annotations

import csv
import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from utils.backup_utils import unwrap_backup_payload
from utils.import_core import as_float, norm_key

MAX_IMPORT_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
MAX_IMPORT_ROWS = 200_000
MAX_CSV_FIELD_SIZE = 1_000_000


@dataclass(frozen=True)
class ParsedImportData:
    path: str
    file_type: str
    rows: list[dict[str, Any]] = field(default_factory=list)
    mandatory_rows: list[dict[str, Any]] = field(default_factory=list)
    wallets: list[dict[str, Any]] = field(default_factory=list)
    initial_balance: float | None = None


def parse_import_file(path: str, *, force: bool = False) -> ParsedImportData:
    source = Path(path)
    if not source.exists():
        raise FileNotFoundError(f"Import file not found: {path}")
    if source.stat().st_size > MAX_IMPORT_FILE_SIZE:
        raise ValueError(f"Import file is too large: {source.stat().st_size} bytes")
    suffix = source.suffix.lower()
    if suffix == ".csv":
        rows = _read_csv_rows(path)
        return ParsedImportData(path=path, file_type="csv", rows=rows)
    if suffix in {".xlsx", ".xlsm"}:
        rows = _read_xlsx_rows(path)
        return ParsedImportData(path=path, file_type="xlsx", rows=rows)
    if suffix == ".json":
        return _read_json_payload(path, force=force)
    raise ValueError(f"Unsupported import file type: {suffix}")


def _normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    return {norm_key(str(k)): v for k, v in row.items() if k is not None}


def _read_csv_rows(path: str) -> list[dict[str, Any]]:
    csv.field_size_limit(MAX_CSV_FIELD_SIZE)
    with open(path, newline="", encoding="utf-8") as csv_file:
        first_pos = csv_file.tell()
        first_line = ""
        while True:
            pos = csv_file.tell()
            line = csv_file.readline()
            if line == "":
                break
            if line.strip():
                first_pos = pos
                first_line = line
                break
        normalized = first_line.lstrip("\ufeff").strip()
        if not normalized.startswith("Transaction statement"):
            csv_file.seek(first_pos)
        reader = csv.DictReader(csv_file)
        rows: list[dict[str, Any]] = []
        for index, row in enumerate(reader, start=1):
            if index > MAX_IMPORT_ROWS:
                raise ValueError(f"CSV import exceeded row limit ({MAX_IMPORT_ROWS})")
            if row:
                rows.append(_normalize_row(row))
        return rows


def _read_xlsx_rows(path: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if not wb.worksheets:
            return []
        ws = wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            return []
        if header_row and str(header_row[0] or "").strip().startswith("Transaction statement"):
            header_row = next(rows_iter, None)
        if header_row is None:
            return []
        headers = [norm_key(str(cell or "")) for cell in header_row]
        rows: list[dict[str, Any]] = []
        for index, row in enumerate(rows_iter, start=1):
            if index > MAX_IMPORT_ROWS:
                raise ValueError(f"XLSX import exceeded row limit ({MAX_IMPORT_ROWS})")
            payload = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            if not any(str(value or "").strip() for value in payload.values()):
                continue
            rows.append(payload)
        return rows
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _read_json_payload(path: str, *, force: bool = False) -> ParsedImportData:
    with open(path, encoding="utf-8") as fp:
        payload = json.load(fp)
    payload = unwrap_backup_payload(payload, force=force)

    wallets = payload.get("wallets", [])
    if not isinstance(wallets, list):
        wallets = []
    records = payload.get("records", [])
    if not isinstance(records, list):
        records = []
    mandatory = payload.get("mandatory_expenses", [])
    if not isinstance(mandatory, list):
        mandatory = []
    transfers = payload.get("transfers", [])
    if not isinstance(transfers, list):
        transfers = []

    rows = [_normalize_row(item) for item in records if isinstance(item, dict)]
    existing_transfer_ids = {
        int(as_float(item.get("transfer_id"), 0.0) or 0)
        for item in rows
        if int(as_float(item.get("transfer_id"), 0.0) or 0) > 0
    }
    rows.extend(_transfer_rows_from_aggregates(transfers, existing_transfer_ids))
    mandatory_rows = []
    for item in mandatory:
        if not isinstance(item, dict):
            continue
        normalized = _normalize_row(item)
        if not str(normalized.get("type", "") or "").strip():
            normalized["type"] = "mandatory_expense"
        mandatory_rows.append(normalized)

    initial_balance = None
    if "initial_balance" in payload:
        initial_balance = as_float(payload.get("initial_balance"), None)
    if initial_balance is None:
        for wallet in wallets:
            if not isinstance(wallet, dict):
                continue
            wallet_id = int(as_float(wallet.get("id"), 0.0) or 0)
            if wallet_id == 1 or bool(wallet.get("system", False)):
                initial_balance = float(as_float(wallet.get("initial_balance"), 0.0) or 0.0)
                break

    return ParsedImportData(
        path=path,
        file_type="json",
        rows=rows,
        mandatory_rows=mandatory_rows,
        wallets=[wallet for wallet in wallets if isinstance(wallet, dict)],
        initial_balance=initial_balance,
    )


def _transfer_rows_from_aggregates(
    items: list[Any], existing_transfer_ids: set[int]
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        transfer_id = int(as_float(item.get("id"), 0.0) or 0)
        if transfer_id > 0 and transfer_id in existing_transfer_ids:
            continue
        rows.append(
            {
                "type": "transfer",
                "date": item.get("date", ""),
                "description": item.get("description", ""),
                "transfer_id": transfer_id if transfer_id > 0 else item.get("id"),
                "amount_original": item.get("amount_original"),
                "currency": item.get("currency", "KZT"),
                "rate_at_operation": item.get("rate_at_operation"),
                "amount_kzt": item.get("amount_kzt"),
                "from_wallet_id": item.get("from_wallet_id"),
                "to_wallet_id": item.get("to_wallet_id"),
            }
        )
    return rows
